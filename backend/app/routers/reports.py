"""
Reports router — generates attendance reports and exports as PDF or Excel.

GET  /api/reports/daily              → daily attendance (filter by date/dept/semester)
GET  /api/reports/monthly            → monthly attendance summary
GET  /api/reports/student/{id}       → per-student full report
GET  /api/reports/department/{id}    → per-department summary
GET  /api/reports/export/excel       → download Excel workbook
GET  /api/reports/export/pdf         → download PDF report
"""
import io
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

from app.core.database import get_db
from app.core.dependencies import require_admin, require_admin_or_faculty
from app.models.attendance_record import AttendanceRecord, AttendanceStatus
from app.models.attendance_session import AttendanceSession
from app.models.department import Department
from app.models.student import Student
from app.models.subject import Subject
from app.schemas.attendance import StudentAttendanceSummary
from app.services.attendance_service import AttendanceService

router = APIRouter(prefix="/reports", tags=["Reports"])


# ── helpers ───────────────────────────────────────────────────────────────────
def _daily_data(
    db: Session,
    report_date: date,
    department_id: Optional[int],
    semester: Optional[int],
    subject_id: Optional[int],
) -> list[dict]:
    """Return raw rows for a daily attendance report."""
    query = (
        db.query(AttendanceRecord)
        .join(AttendanceRecord.session)
        .join(AttendanceRecord.student)
        .options(
            joinedload(AttendanceRecord.student),
            joinedload(AttendanceRecord.session).joinedload(AttendanceSession.subject),
            joinedload(AttendanceRecord.session).joinedload(AttendanceSession.department),
        )
        .filter(AttendanceSession.session_date == report_date)
    )
    if department_id:
        query = query.filter(AttendanceSession.department_id == department_id)
    if semester:
        query = query.filter(AttendanceSession.semester == semester)
    if subject_id:
        query = query.filter(AttendanceSession.subject_id == subject_id)

    rows = []
    for record in query.all():
        rows.append({
            "date": str(record.session.session_date),
            "session_id": record.session_id,
            "subject": record.session.subject.subject_name if record.session.subject else "",
            "department": record.session.department.name if record.session.department else "",
            "semester": record.session.semester,
            "enrollment_no": record.student.enrollment_no if record.student else "",
            "student_name": record.student.full_name if record.student else "",
            "status": record.status.value,
            "marked_at": str(record.marked_at),
        })
    return rows


# ── Endpoints ─────────────────────────────────────────────────────────────────
@router.get("/daily", summary="Daily attendance report")
def daily_report(
    report_date: date = Query(default=date.today()),
    department_id: Optional[int] = Query(None),
    semester: Optional[int] = Query(None, ge=1, le=8),
    subject_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_faculty),
):
    rows = _daily_data(db, report_date, department_id, semester, subject_id)
    return {"date": str(report_date), "total_records": len(rows), "records": rows}


@router.get("/monthly", summary="Monthly attendance summary by student")
def monthly_report(
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    department_id: Optional[int] = Query(None),
    semester: Optional[int] = Query(None, ge=1, le=8),
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_faculty),
):
    from calendar import monthrange

    _, last_day = monthrange(year, month)
    date_from = date(year, month, 1)
    date_to = date(year, month, last_day)

    student_q = db.query(Student).filter(Student.status == True)  # noqa: E712
    if department_id:
        student_q = student_q.filter(Student.department_id == department_id)
    if semester:
        student_q = student_q.filter(Student.semester == semester)

    svc = AttendanceService(db)
    summaries = []
    for student in student_q.all():
        summary = svc.get_student_summary(student.id, date_from=date_from, date_to=date_to)
        summaries.append(summary.model_dump())

    return {
        "year": year,
        "month": month,
        "total_students": len(summaries),
        "summaries": summaries,
    }


@router.get("/student/{student_id}", response_model=StudentAttendanceSummary, summary="Student attendance report")
def student_report(
    student_id: int,
    subject_id: Optional[int] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_faculty),
):
    return AttendanceService(db).get_student_summary(
        student_id=student_id,
        subject_id=subject_id,
        date_from=date_from,
        date_to=date_to,
    )


@router.get("/department/{department_id}", summary="Department attendance report")
def department_report(
    department_id: int,
    semester: Optional[int] = Query(None, ge=1, le=8),
    subject_id: Optional[int] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_faculty),
):
    summaries = AttendanceService(db).get_department_report(
        department_id=department_id,
        semester=semester,
        subject_id=subject_id,
        date_from=date_from,
        date_to=date_to,
    )
    dept = db.query(Department).filter(Department.id == department_id).first()
    return {
        "department": dept.name if dept else str(department_id),
        "total_students": len(summaries),
        "summaries": [s.model_dump() for s in summaries],
    }


# ── Export: Excel ─────────────────────────────────────────────────────────────
@router.get("/export/excel", summary="Export attendance report as Excel (.xlsx)")
def export_excel(
    report_date: date = Query(default=date.today()),
    department_id: Optional[int] = Query(None),
    semester: Optional[int] = Query(None, ge=1, le=8),
    subject_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_faculty),
):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    rows = _daily_data(db, report_date, department_id, semester, subject_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # ── Header styling ────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    headers = [
        "Date", "Session ID", "Department", "Semester",
        "Subject", "Enrollment No", "Student Name", "Status", "Marked At",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ── Data rows ─────────────────────────────────────────────────
    present_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    absent_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")

    for row_idx, row in enumerate(rows, start=2):
        values = [
            row["date"], row["session_id"], row["department"], row["semester"],
            row["subject"], row["enrollment_no"], row["student_name"],
            row["status"], row["marked_at"],
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if col == 8:  # Status column
                cell.fill = present_fill if val == "PRESENT" else absent_fill

    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"attendance_{report_date}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Export: PDF ───────────────────────────────────────────────────────────────
@router.get("/export/pdf", summary="Export attendance report as PDF")
def export_pdf(
    report_date: date = Query(default=date.today()),
    department_id: Optional[int] = Query(None),
    semester: Optional[int] = Query(None, ge=1, le=8),
    subject_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_faculty),
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    rows = _daily_data(db, report_date, department_id, semester, subject_id)

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("Attendance Report", styles["Title"]))
    elements.append(Paragraph(f"Date: {report_date}  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 0.5 * cm))

    # Table data
    col_headers = ["#", "Enrollment No", "Student Name", "Department", "Sem", "Subject", "Status"]
    table_data = [col_headers]
    for i, row in enumerate(rows, start=1):
        table_data.append([
            str(i),
            row["enrollment_no"],
            row["student_name"],
            row["department"],
            str(row["semester"]),
            row["subject"],
            row["status"],
        ])

    col_widths = [1 * cm, 3 * cm, 5 * cm, 4 * cm, 1.5 * cm, 5 * cm, 2.5 * cm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF3FB")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        # Colour-code PRESENT / ABSENT
        *[
            ("BACKGROUND", (6, i + 1), (6, i + 1),
             colors.HexColor("#C6EFCE") if row["status"] == "PRESENT" else colors.HexColor("#FFC7CE"))
            for i, row in enumerate(rows)
        ],
    ]))
    elements.append(table)

    doc.build(elements)
    output.seek(0)

    filename = f"attendance_{report_date}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
